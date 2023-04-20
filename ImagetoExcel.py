import xlsxwriter
from PIL import Image
from openpyxl import load_workbook
from pathlib import Path

def ReadExcelRoot(pathExcelRoot):

    wb = load_workbook(filename = pathExcelRoot)
    # WBsheetData = wb['label']
    WBsheetData = wb.active
    
    # Total Row in Excel
    maxRow = WBsheetData.max_row+1

    # Ignore Fist Row because It is Header/Title
    start = 2
    while (start <= maxRow):
        # listCount.append(f"{(WBsheetData.cell(row=start, column=1).value)} | {(WBsheetData.cell(row=start, column=2).value)}")
        TextCount = (f"{(WBsheetData.cell(row=start, column=1).value)}|||{(WBsheetData.cell(row=start, column=2).value)}")
        
        yield TextCount
        start += 1

def ExportImagesToExcel(pathExcelRoot, folderListImages, pathExcelWrite, folderExcelResult):
    # Create an new Excel file and add a worksheet
    workbook = xlsxwriter.Workbook(f"{folderExcelResult}/{pathExcelWrite}")
    worksheet = workbook.add_worksheet()

    # Height Image and Cell
    cell_height = 30.0

    # Set Default Row
    worksheet.set_default_row(int(cell_height))
    worksheet.set_column(1,1, 50)

    # Texts in Excel Root
    TextCount = ReadExcelRoot(pathExcelRoot)

    numberA = numberB = numberC = 1
    try:
        for T in TextCount:
            TextToFile = T.split('|||')
            try:
                # Insert Text to Rows a.k.a Column A
                worksheet.write('A'+str(numberA), TextToFile[0])

                # Parameters Image to Row Excel
                filename = f"{folderListImages}/{TextToFile[0]}.png"
                img = Image.open(filename)
                image_width, image_height = img.size
                y_scale = cell_height/image_height
                cellImage = 'B' + str(numberB)
                # Insert Image to Row Excel
                worksheet.insert_image(cellImage, filename, {'x_scale': y_scale, 'y_scale': y_scale})

                # Insert Text to Rows a.k.a Column C
                worksheet.write('C'+str(numberC), TextToFile[1])
                
                numberC += 1
                numberA += 1
                numberB += 1

            except FileNotFoundError:
                continue
    except KeyboardInterrupt:
        print(KeyboardInterrupt)
        workbook.close()
    workbook.close()




if __name__ == "__main__":
    pathExcelRoot = "Label_CMT_20_03_2021_Real"+".xlsx"
    folderListImages = '31'
    pathExcelWrite = "ImageToExcel"+".xlsx"
    
    # Create folder contain Excels safely
    folderExcelResult = "ExportExcelFolder"
    Path(folderExcelResult).mkdir(parents=True, exist_ok=True)
    
    """
    pathExcelRoot = 'images - Copy.xlsx'
    folderListImages = 'imagesDATA'
    pathExcelWrite = 'ImagesToExcel.xlsx'
    """

    # ===>> Uncomment and Run here <<====
    ExportImagesToExcel(pathExcelRoot, folderListImages, pathExcelWrite, folderExcelResult)
